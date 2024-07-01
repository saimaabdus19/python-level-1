import openpyxl
import matplotlib.pyplot as plt

def create_excel_file(filename):
    # Create a new excel file and add some data
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Data"
  
    # Add headers
    headers = ["Months", "Sales"]
    sheet.append(headers)

    # Add sample data
    data = [
        ["January", 150],
        ["February", 200],
        ["March", 250],
        ["April", 300],
        ["May", 350],
        ["June", 320],
        ["July", 700],
        ["August", 210],
        ["September", 160],
        ["October", 205],
        ["November", 500],
        ["December", 340]
    ]
    
    for row in data:
        sheet.append(row)
  
    workbook.save(filename)
    print(f"The excel file {filename} was created successfully!")
  
def read_excel_file(filename):
    workbook = openpyxl.load_workbook(filename)
    sheet = workbook["Data"]  # Correct sheet name

    data = []
    for row in sheet.iter_rows(values_only=True):  # It will tell it to just go through when there is actual data 
        data.append(row)
  
    return data

def visualise(data):
    # Extract months and sales data
    months = [row[0] for row in data[1:]]  # Exclude header
    sales = [row[1] for row in data[1:]]   # Exclude header
  
    plt.figure(figsize=(10, 6))
    plt.plot(months, sales, marker='o', linestyle='-', color='b')
    plt.title('Monthly Sales Data')
    plt.xlabel('Month')
    plt.ylabel('Sales')
    plt.grid(True)
    plt.show()
  
def main():
    filename = 'sales_data.xlsx'
  
    # Create Excel File
    create_excel_file(filename)
  
    # Read data from Excel file
    data = read_excel_file(filename)
    print("Data from excel file: ")
    for row in data:
        print(row)
    
    # Visualize the data in a line graph
    visualise(data)
  
if __name__ == "__main__":
    main()
