import openpyxl
import matplotlib.pyplot as plt
from openpyxl.utils import get_column_letter
import os

def create_excel_file(filename):
    # Create a new excel file and add headers
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Attendance"
    
    # Add headers
    headers = ["Employee Name", "Date", "Status"]
    sheet.append(headers)
    
    workbook.save(filename)
    print(f"The excel file {filename} was created successfully!")

def record_attendance(filename):
    # Load the workbook and select the active sheet
    workbook = openpyxl.load_workbook(filename)
    sheet = workbook.active
    
    # Prompt the user to enter attendance details
    employee_name = input("Enter employee name: ")
    date = input("Enter date (YYYY-MM-DD): ")
    status = input("Enter status (Present, Absent, Late): ")
    
    # Append the details to the sheet
    sheet.append([employee_name, date, status])
    
    workbook.save(filename)
    print("Attendance recorded successfully.")

def view_attendance_records(filename):
    # Load the workbook and select the active sheet
    workbook = openpyxl.load_workbook(filename)
    sheet = workbook.active
    
    # Read the attendance records
    data = []
    for row in sheet.iter_rows(values_only=True):
        data.append(row)
    
    # Display the attendance records
    print("Displaying Attendance Records:")
    for row in data:
        print(", ".join(row))

def visualize_attendance(filename):
    # Load the workbook and select the active sheet
    workbook = openpyxl.load_workbook(filename)
    sheet = workbook.active
    
    # Read the attendance records
    data = []
    for row in sheet.iter_rows(values_only=True):
        data.append(row)
    
    # Extract dates and status for visualization
    dates = [row[1] for row in data[1:]]  # Exclude header
    status = [row[2] for row in data[1:]] # Exclude header
    
    # Count occurrences of each status
    status_counts = {"Present": 0, "Absent": 0, "Late": 0}
    for s in status:
        if s in status_counts:
            status_counts[s] += 1
    
    # Create a bar chart for attendance status counts
    plt.figure(figsize=(10, 6))
    plt.bar(status_counts.keys(), status_counts.values(), color=['green', 'red', 'orange'])
    plt.title('Attendance Status Counts')
    plt.xlabel('Status')
    plt.ylabel('Count')
    plt.grid(True)
    plt.show()

def main():
    filename = 'attendance.xlsx'
    
    # Check if the Excel file exists; if not, create it
    if not os.path.exists(filename):
        create_excel_file(filename)
    
    print("Welcome to the Attendance Tracker!")
    
    while True:
        print("\nPlease choose an option:")
        print("1. Record attendance")
        print("2. View all attendance records")
        print("3. Visualize attendance trends")
        print("4. Quit")
        
        try:
            choice = int(input("Enter your choice: "))
        except ValueError:
            print("Invalid input. Please enter a number (1, 2, 3, or 4).")
            continue
        
        if choice == 1:
            record_attendance(filename)
        elif choice == 2:
            view_attendance_records(filename)
        elif choice == 3:
            visualize_attendance(filename)
        elif choice == 4:
            print("Thank you for using the Attendance Tracker. Goodbye!")
            break
        else:
            print("Invalid option. Please enter 1, 2, 3, or 4.")

if __name__ == "__main__":
    main()
